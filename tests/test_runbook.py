"""Runbook: auth gate, catalogue, read-only enforcement, runners, export."""

from __future__ import annotations

import io

import pytest

from tests.conftest import RUNBOOK_PASSWORD, RUNBOOK_USER


# --- Auth -------------------------------------------------------------------


def test_runbook_requires_login(client):
    client.cookies.clear()
    body = client.get("/runbook").text
    assert "Sign in to the Runbook" in body
    # The catalogue must not leak before sign-in.
    assert "Host Inventory Backup" not in body


def test_wrong_password_is_rejected(client):
    client.cookies.clear()
    resp = client.post(
        "/runbook/login", data={"username": RUNBOOK_USER, "password": "nope"}
    )
    assert resp.status_code == 401
    assert "Wrong username or password" in resp.text
    assert not resp.cookies.get("samix_runbook")


def test_unknown_user_is_rejected(client):
    client.cookies.clear()
    resp = client.post(
        "/runbook/login", data={"username": "mallory", "password": RUNBOOK_PASSWORD}
    )
    assert resp.status_code == 401


def test_login_then_catalogue_and_logout(client):
    client.cookies.clear()
    resp = client.post(
        "/runbook/login",
        data={"username": RUNBOOK_USER, "password": RUNBOOK_PASSWORD},
        follow_redirects=True,
    )
    assert resp.status_code == 200
    assert "Host Inventory Backup" in resp.text
    assert f"Signed in as" in resp.text

    # Signing out clears the cookie and puts the gate back.
    client.post("/runbook/logout", follow_redirects=True)
    assert "Sign in to the Runbook" in client.get("/runbook").text


def test_tampered_cookie_is_not_accepted(client):
    client.cookies.clear()
    client.cookies.set("samix_runbook", "admin|99999999999|forged-signature")
    assert "Sign in to the Runbook" in client.get("/runbook").text


def test_expired_token_is_rejected():
    from app.config import get_settings
    from app.runbook_auth import read_token

    settings = get_settings()
    assert read_token(settings, "admin|1|deadbeef") is None


def test_token_roundtrip_and_digest():
    from app.config import get_settings
    from app.runbook_auth import issue_token, password_digest, read_token, verify_credentials

    settings = get_settings()
    assert read_token(settings, issue_token(settings, "admin")) == "admin"
    assert verify_credentials(settings, "ADMIN", RUNBOOK_PASSWORD) == "admin"
    assert verify_credentials(settings, "admin", RUNBOOK_PASSWORD + "x") is None
    # The configured value really is a digest, not the password.
    assert password_digest(RUNBOOK_PASSWORD) != RUNBOOK_PASSWORD


# --- Catalogue --------------------------------------------------------------


@pytest.fixture
def admin(client):
    """A client signed in to the Runbook."""
    client.cookies.clear()
    client.post(
        "/runbook/login",
        data={"username": RUNBOOK_USER, "password": RUNBOOK_PASSWORD},
    )
    return client


def test_platform_filter_narrows_the_catalogue(admin):
    from app.runbook import SCRIPTS

    zbx = admin.get("/runbook?platform=zabbix").text
    assert "Proxy Status" in zbx
    # Every Zabbix script is listed; nothing from another platform sneaks in.
    for s in SCRIPTS:
        if s.platform == "zabbix":
            assert s.title in zbx

    nnmi = admin.get("/runbook?platform=nnmi").text
    assert "Proxy Status" not in nnmi


def test_script_page_shows_full_documentation(admin):
    body = admin.get("/runbook/ip-monitoring-status").text
    assert "What it does" in body and "Step by step" in body
    # Declared columns are all advertised on the page.
    from app.runbook import SCRIPTS_BY_SLUG

    for col in SCRIPTS_BY_SLUG["ip-monitoring-status"].columns:
        assert col in body
    assert "host.get" in body and "item.get" in body


def test_unknown_script_is_a_404(admin):
    assert admin.get("/runbook/does-not-exist").status_code == 404


# --- Read-only enforcement --------------------------------------------------


def test_write_scripts_are_documented_but_have_no_runner():
    from app.runbook import SCRIPTS

    writers = [s for s in SCRIPTS if not s.read_only]
    assert {s.slug for s in writers} == {"add-hosts", "add-users", "disable-hosts"}
    for s in writers:
        assert s.runner is None, f"{s.slug} must not be runnable from the web"
        assert s.purpose and s.steps and s.notes  # still fully documented


def test_write_script_page_has_no_run_button(admin):
    body = admin.get("/runbook/disable-hosts").text
    assert "changes the monitoring configuration" in body
    assert "Run script" not in body


def test_executing_a_write_script_is_refused():
    from app.config import get_settings
    from app.runbook import SCRIPTS_BY_SLUG, RunbookError, execute

    with pytest.raises(RunbookError, match="cannot be run"):
        execute(SCRIPTS_BY_SLUG["disable-hosts"], "all", {}, get_settings())


def test_read_rpc_refuses_non_get_methods():
    from app.collectors.base import CollectorError
    from app.collectors.zabbix import ZabbixCollector
    from app.config import get_settings
    from app.servers import ServerConfig

    c = ZabbixCollector(
        ServerConfig(name="Z", platform="zabbix", url="http://x"), get_settings()
    )
    for method in ("host.create", "host.update", "user.create", "host.massupdate"):
        with pytest.raises(CollectorError, match="refuses non-read"):
            c.read_rpc(method, {})


# --- Runners (against a fake collector, no network) -------------------------


class FakeZabbix:
    """Stands in for a live collector: records calls, replays canned results."""

    name = "zabbix"

    def __init__(self, instance="Zabbix-TEST", **responses):
        self.instance = instance
        self.calls: list[tuple[str, dict]] = []
        self._responses = responses

        class _S:
            mock_mode = False
            runbook_max_rows = 20000

        self.settings = _S()

    def read_rpc(self, method, params):
        self.calls.append((method, params))
        value = self._responses.get(method, [])
        return value(params) if callable(value) else value


def _host(hostid, name, ip, status="0", available="0", groups=("Prod",),
          templates=("Linux by Zabbix agent",)):
    return {
        "hostid": hostid, "host": name, "name": name, "status": status,
        "interfaces": [{"ip": ip, "dns": "", "useip": "1", "main": "1",
                        "available": available, "error": "agent timed out"}],
        "hostgroups": [{"name": g} for g in groups],
        "parentTemplates": [{"name": t} for t in templates],
    }


def test_unavailable_hosts_reads_the_interface_not_the_host():
    from app.runbook import run_unavailable_hosts

    fake = FakeZabbix(**{"host.get": [
        _host("1", "web-01", "10.0.0.1", available="2"),
        _host("2", "web-02", "10.0.0.2", available="1"),
    ]})
    rows = run_unavailable_hosts([fake], {})
    assert [r[1] for r in rows] == ["web-01"]
    assert rows[0][6] == "agent timed out"  # the interface's own error
    # Only enabled hosts are asked for — a deliberately-disabled host is not a fault.
    assert fake.calls[0][1]["filter"] == {"status": "0"}


def test_ip_lookup_filters_server_side_and_reports_misses():
    from app.runbook import run_ip_lookup

    fake = FakeZabbix(**{"host.get": [_host("7", "db-01", "10.0.0.7")]})
    rows = run_ip_lookup([fake], {"ips": "10.0.0.7, 10.0.0.9"})

    # The IP filter goes to Zabbix rather than being applied in Python.
    method, params = fake.calls[0]
    assert method == "host.get"
    assert params["filter"]["ip"] == ["10.0.0.7", "10.0.0.9"]

    by_ip = {r[0]: r for r in rows}
    assert by_ip["10.0.0.7"][3] == "db-01"
    assert by_ip["10.0.0.9"][5] == "NOT FOUND"


def test_ip_lookup_needs_at_least_one_ip():
    from app.runbook import RunbookError, run_ip_lookup

    with pytest.raises(RunbookError, match="at least one IP"):
        run_ip_lookup([FakeZabbix()], {"ips": "   "})


def test_monitoring_status_counts_items_in_one_batched_call():
    from app.runbook import run_ip_monitoring_status

    fake = FakeZabbix(**{
        "host.get": [_host("1", "a", "10.0.0.1"), _host("2", "b", "10.0.0.2")],
        # host 1 has two items, host 2 has none
        "item.get": [{"hostid": "1"}, {"hostid": "1"}],
    })
    rows = run_ip_monitoring_status([fake], {"ips": "10.0.0.1 10.0.0.2"})
    by_ip = {r[0]: r for r in rows}
    assert by_ip["10.0.0.1"][6] == "MONITORED" and by_ip["10.0.0.1"][7] == 2
    # Enabled but collecting nothing is NOT monitored — the whole point.
    assert by_ip["10.0.0.2"][6] == "NOT MONITORED" and by_ip["10.0.0.2"][7] == 0

    item_calls = [c for c in fake.calls if c[0] == "item.get"]
    assert len(item_calls) == 1, "items must be counted for all hosts in one call"
    assert sorted(item_calls[0][1]["hostids"]) == ["1", "2"]


def test_monitoring_status_marks_disabled_hosts_not_monitored():
    from app.runbook import run_ip_monitoring_status

    fake = FakeZabbix(**{
        "host.get": [_host("1", "old", "10.0.0.1", status="1")],
        "item.get": [{"hostid": "1"}] * 5,
    })
    row = run_ip_monitoring_status([fake], {"ips": "10.0.0.1"})[0]
    assert row[5] == "Disabled" and row[6] == "NOT MONITORED"


def test_proxy_status_handles_both_zabbix_schemas():
    from app.runbook import run_proxy_status

    modern = FakeZabbix("Z-New", **{"proxy.get": [{
        "proxyid": "1", "name": "prx-a", "state": "2", "address": "10.0.0.5",
        "port": "10051", "operating_mode": "1", "lastaccess": "1700000000",
        "version": "70207", "hosts": 42,
    }]})
    legacy = FakeZabbix("Z-Old", **{"proxy.get": [{
        "proxyid": "9", "host": "prx-legacy", "status": "5",
        "lastaccess": "1", "hosts": 3,
    }]})
    rows = {r[2]: r for r in run_proxy_status([modern, legacy], {})}

    assert rows["prx-a"][3] == "ONLINE" and rows["prx-a"][6] == "PASSIVE"
    assert rows["prx-a"][9] == 42
    # Legacy row: name from `host`, mode from `status`, state inferred from age.
    assert rows["prx-legacy"][3] == "OFFLINE" and rows["prx-legacy"][6] == "ACTIVE"


def test_group_audit_batches_item_counts_and_labels_groups():
    from app.runbook import run_group_audit

    fake = FakeZabbix(**{
        "hostgroup.get": [{"groupid": "10", "name": "Vcenter_10.1.1.1"}],
        "host.get": [{
            "hostid": "5", "host": "esx-1", "name": "esx-1", "status": "0",
            "hostgroups": [{"groupid": "10", "name": "Vcenter_10.1.1.1"}],
            "parentTemplates": [{"name": "VMware"}],
        }],
        "item.get": [{"hostid": "5"}] * 3,
    })
    rows = run_group_audit([fake], {"group": "Vcenter_"})
    assert rows[0][1] == "Vcenter_10.1.1.1"
    assert rows[0][6] == "MONITORED" and rows[0][7] == 3
    assert len([c for c in fake.calls if c[0] == "item.get"]) == 1


def test_group_audit_explains_an_empty_search():
    from app.runbook import RunbookError, run_group_audit

    with pytest.raises(RunbookError, match="No host group matched"):
        run_group_audit([FakeZabbix()], {"group": "Nothing"})


def test_host_backup_flattens_tags_and_inventory_mode():
    from app.runbook import run_host_backup

    host = _host("3", "app-01", "10.0.0.3")
    host["inventory_mode"] = "1"
    host["tags"] = [{"tag": "env", "value": "prod"}, {"tag": "critical", "value": ""}]
    rows = run_host_backup([FakeZabbix(**{"host.get": [host]})], {})
    assert rows[0][8] == "env=prod; critical"
    assert rows[0][9] == "Automatic"


def test_ip_history_groups_calls_by_value_type():
    from app.runbook import run_ip_history

    fake = FakeZabbix(**{
        "host.get": [_host("1", "srv", "10.0.0.1")],
        "item.get": [
            {"itemid": "11", "name": "CPU", "key_": "cpu", "units": "%", "value_type": "0"},
            {"itemid": "12", "name": "Load", "key_": "load", "units": "", "value_type": "0"},
            {"itemid": "13", "name": "Uptime", "key_": "up", "units": "s", "value_type": "3"},
            {"itemid": "14", "name": "Log", "key_": "log", "units": "", "value_type": "2"},
        ],
        "history.get": lambda p: [
            {"itemid": p["itemids"][0], "clock": "1700000000", "value": "1.5"}
        ],
    })
    rows = run_ip_history([fake], {
        "ips": "10.0.0.1",
        "date_from": "2026-08-01T00:00",
        "date_to": "2026-08-02T00:00",
    })
    history_calls = [c for c in fake.calls if c[0] == "history.get"]
    # Two numeric value types -> two calls, not one per item, and text is skipped.
    assert len(history_calls) == 2
    assert {c[1]["history"] for c in history_calls} == {0, 3}
    assert sorted(history_calls[0][1]["itemids"]) == ["11", "12"]
    assert rows[0][4] == "CPU"


def test_ip_history_rejects_a_backwards_window():
    from app.runbook import RunbookError, run_ip_history

    with pytest.raises(RunbookError, match="later than"):
        run_ip_history([FakeZabbix()], {
            "ips": "10.0.0.1",
            "date_from": "2026-08-02T00:00",
            "date_to": "2026-08-01T00:00",
        })


def test_ip_history_wants_exactly_one_ip():
    from app.runbook import RunbookError, run_ip_history

    with pytest.raises(RunbookError, match="exactly one IP"):
        run_ip_history([FakeZabbix()], {"ips": "10.0.0.1 10.0.0.2"})


def test_runners_explain_mock_mode_instead_of_failing_obscurely():
    from app.runbook import RunbookError, _require_zabbix

    fake = FakeZabbix()
    fake.settings.mock_mode = True
    with pytest.raises(RunbookError, match="MOCK_MODE"):
        _require_zabbix([fake])


def test_no_instance_selected_is_explained():
    from app.runbook import RunbookError, _require_zabbix

    with pytest.raises(RunbookError, match="No Zabbix instance"):
        _require_zabbix([])


# --- Run + export end to end ------------------------------------------------


def test_run_endpoint_renders_results_and_export_carries_the_credit(admin, monkeypatch):
    import app.runbook as rb

    fake = FakeZabbix(**{"host.get": [_host("1", "web-01", "10.0.0.1")]})
    monkeypatch.setattr(rb, "collectors_for", lambda instance, platform: [fake])

    resp = admin.post(
        "/partials/runbook/run/ip-lookup",
        data={"instance": "all", "ips": "10.0.0.1"},
    )
    assert resp.status_code == 200
    assert "web-01" in resp.text and "Export Excel" in resp.text

    token = resp.text.split("token=")[1].split('"')[0]
    xlsx = admin.get(f"/runbook/ip-lookup/export.xlsx?instance=all&token={token}")
    assert xlsx.status_code == 200
    assert "SAMIX_ip-lookup" in xlsx.headers["content-disposition"]

    from openpyxl import load_workbook

    ws = load_workbook(io.BytesIO(xlsx.content)).active
    text = "\n".join(
        str(c.value) for row in ws.iter_rows(max_row=8) for c in row if c.value
    )
    assert "Script by Eng. Ahmed Hussien" in text
    assert "IP" in text and "web-01" in text


def test_run_endpoint_surfaces_errors_instead_of_500ing(admin):
    resp = admin.post(
        "/partials/runbook/run/ip-lookup", data={"instance": "all", "ips": ""}
    )
    assert resp.status_code == 200
    assert "IP addresses is required" in resp.text


def test_run_endpoint_needs_a_session(client):
    client.cookies.clear()
    resp = client.post(
        "/partials/runbook/run/ip-lookup", data={"instance": "all", "ips": "10.0.0.1"}
    )
    assert resp.status_code == 401


def test_export_needs_a_session(client):
    client.cookies.clear()
    assert client.get("/runbook/ip-lookup/export.xlsx").status_code == 401


# --- Catalogue integrity ----------------------------------------------------


def test_every_script_is_fully_documented():
    from app.runbook import SCRIPTS

    slugs = [s.slug for s in SCRIPTS]
    assert len(slugs) == len(set(slugs)), "slugs must be unique"
    for s in SCRIPTS:
        assert s.tagline and s.purpose and s.steps and s.columns and s.api_calls
        assert s.platform in ("zabbix", "dynatrace", "nnmi")
        assert s.author == "Eng. Ahmed Hussien"
        if s.read_only:
            assert s.runner is not None, f"{s.slug} is read-only but has no runner"


def test_runner_row_width_matches_declared_columns():
    """A runner returning the wrong number of cells would silently skew the export."""
    from app.runbook import SCRIPTS_BY_SLUG, run_disabled_hosts, run_unavailable_hosts

    fake = FakeZabbix(**{"host.get": [_host("1", "h", "10.0.0.1", status="1",
                                            available="2")]})
    for slug, runner in (
        ("disabled-hosts", run_disabled_hosts),
        ("unavailable-hosts", run_unavailable_hosts),
    ):
        rows = runner([fake], {})
        assert rows, slug
        assert len(rows[0]) == len(SCRIPTS_BY_SLUG[slug].columns), slug
