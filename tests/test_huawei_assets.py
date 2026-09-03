"""Huawei i2000 / Digital View asset inventory import."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from app.huawei_assets import parse_workbook
from app.models import HostStatus


def _template(tmp_path: Path) -> Path:
    """A miniature BaseAssetImportTemplate with the real sheet geometry.

    Rows 1-5 are Huawei's metadata banner, row 6 the real headers, row 7 a row
    of internal field ids, and the assets start at row 8. The detail sheets are
    shallower: headers on row 3, field ids on row 4, data from row 5.
    """
    wb = Workbook()
    wb.remove(wb.active)

    vm = wb.create_sheet("VM Operating System")
    for _ in range(5):
        vm.append([])
    vm.append(["*Name", "IP", "AssetVersion", "Vendor", "Administrator",
               "Location", "Description", "Site name", "Host Name", "OS Type",
               "CPU", "Memory(MB)", "Application type"])
    vm.append(["", "", "", "", "", "", "", "SiteName", "HostName", "OSType",
               "CPU", "Memory", "InstallApp"])          # field-id row
    vm.append(["vm-app-01", "10.9.0.1", "SUSE12", "huawei", "", "", "",
               "Site-PR", "vm-app-01", "Linux", 8, 32768, "AG"])
    vm.append(["vm-app-02", "10.9.0.2", "SUSE12", "huawei", "", "", "",
               "Site-DR", "vm-app-02", "Linux", 16, 65536, "BCS"])

    pm = wb.create_sheet("PM Operating System")
    for _ in range(5):
        pm.append([])
    pm.append(["*Name", "IP", "AssetVersion", "Vendor", "Administrator",
               "Location", "Description", "Site name", "Host Name", "OS Type",
               "CPU", "Memory(MB)"])
    pm.append(["", "", "", "", "", "", "", "SiteName", "HostName", "OSType",
               "CPU", "Memory"])
    pm.append(["pm-db-01", "10.9.1.1", "/", "huawei", "", "C1S14", "", "Site-PR",
               "pm-db-01", "Linux", 64, 1048576])

    rack = wb.create_sheet("Rack Server")
    for _ in range(5):
        rack.append([])
    rack.append(["*Name", "BMC IP", "AssetVersion", "Vendor", "Administrator",
                 "Location", "Description", "Site name", "*Device type",
                 "Model No.", "Sequence No.", "User", "Password"])
    rack.append(["", "", "", "", "", "", "", "SiteName", "DeviceType",
                 "Model", "Serial", "User", "Password"])
    rack.append(["rack-01", "10.9.2.1", "", "huawei", "", "C1S26", "",
                 "Site-DR", "2288H", "2288H V6", "SN123", "Administrator",
                 "SUPERSECRET"])

    disks = wb.create_sheet("Disk Information")
    disks.append([]); disks.append([])
    disks.append(["*TypeGroupID", "*TypeID", "*Name", "*Disk path", "*Capacity(GB)"])
    disks.append(["", "", "", "Device", "Size"])
    disks.append(["I2000Host", "VMOS", "vm-app-01", "/", 16])
    disks.append(["I2000Host", "VMOS", "vm-app-01", "/data", 500])
    disks.append(["I2000Host", "PMOS", "pm-db-01", "/", 100])

    vlan = wb.create_sheet("VLAN Information")
    vlan.append([]); vlan.append([])
    vlan.append(["*TypeGroupID", "*TypeID", "*Name", "Network ID", "VLAN ID",
                 "*IP address", "MAC address"])
    vlan.append(["", "", "", "VLANID", "SEGMENTATIONID", "IP", "MAC"])
    vlan.append(["I2000Host", "VMOS", "vm-app-01", "", "", "10.9.0.1", "FA:16:01"])
    vlan.append(["I2000Host", "VMOS", "vm-app-01", "", "", "10.9.90.1", "FA:16:02"])

    comp = wb.create_sheet("Component Information")
    comp.append([]); comp.append([])
    comp.append(["*TypeGroupID", "*TypeID", "*Name", "Component Type",
                 "*Application unit", "Deployment mode", "Deployment path"])
    comp.append(["", "", "", "ComponentName", "Appunit", "DeployType", "DeployPath"])
    comp.append(["I2000Host", "VMOS", "vm-app-01", "AG_CLUSTER", "AG", "CLUSTER", ""])

    acct = wb.create_sheet("Account Information")
    acct.append([]); acct.append([])
    acct.append(["*TypeGroupID", "*TypeID", "*Name", "Account type",
                 "*Protocol type", "Port", "*Account name", "*Password"])
    acct.append(["", "", "", "AccountType", "Protocol", "Port", "UserName", "Password"])
    acct.append(["I2000Host", "VMOS", "vm-app-01", "", "SSH", "22", "root", "hunter2"])

    path = tmp_path / "BaseAssetImportTemplate_En.xlsx"
    wb.save(path)
    return path


@pytest.fixture(scope="module")
def inventory(tmp_path_factory):
    return parse_workbook(_template(tmp_path_factory.mktemp("huawei")))


# --- Parsing ----------------------------------------------------------------


def test_reads_every_asset_sheet(inventory):
    assert inventory.count == 4
    assert inventory.sheets_read["VM Operating System"] == 2
    assert inventory.sheets_read["PM Operating System"] == 1
    assert inventory.sheets_read["Rack Server"] == 1


def test_field_id_row_is_not_imported_as_an_asset(inventory):
    """Row 7 holds Huawei's internal field ids, not a machine."""
    names = {h["hostname"] for h in inventory.hosts}
    assert "SiteName" not in names and "HostName" not in names
    assert names == {"vm-app-01", "vm-app-02", "pm-db-01", "rack-01"}


def test_assets_are_keyed_by_name_and_site(inventory):
    """The same name exists at both the PR and DR site; the key must separate them."""
    ids = {h["external_id"] for h in inventory.hosts}
    assert "vm-app-01@Site-PR" in ids
    assert "vm-app-02@Site-DR" in ids


def test_capacity_comes_across(inventory):
    by_name = {h["hostname"]: h for h in inventory.hosts}
    vm = by_name["vm-app-01"]["metrics"]
    assert vm["cores"] == 8
    assert vm["mem_total_gb"] == 32.0        # 32768 MB
    assert vm["disk_total_gb"] == 516.0      # 16 + 500, summed across partitions

    pm = by_name["pm-db-01"]["metrics"]
    assert pm["cores"] == 64 and pm["mem_total_gb"] == 1024.0


def test_application_unit_becomes_the_group(inventory):
    """AG / BCS is what an operator groups by, not the site."""
    by_name = {h["hostname"]: h for h in inventory.hosts}
    assert by_name["vm-app-01"]["group_name"] == "AG"
    assert by_name["vm-app-02"]["group_name"] == "BCS"
    # A rack server has no application unit; it falls back to its site.
    assert by_name["rack-01"]["group_name"] == "Site-DR"


def test_details_are_attached(inventory):
    payload = next(
        h["raw_payload"] for h in inventory.hosts if h["hostname"] == "vm-app-01"
    )
    assert payload["components"] == ["AG_CLUSTER"]
    assert sorted(payload["all_ips"]) == ["10.9.0.1", "10.9.90.1"]
    assert payload["os_type"] == "Linux"


def test_rack_server_uses_its_bmc_ip_and_keeps_hardware_detail(inventory):
    rack = next(h for h in inventory.hosts if h["hostname"] == "rack-01")
    assert rack["ip"] == "10.9.2.1"
    assert rack["raw_payload"]["model"] == "2288H V6"
    assert rack["raw_payload"]["serial"] == "SN123"
    assert rack["raw_payload"]["rack_location"] == "C1S26"


# --- Inventory, not monitoring ----------------------------------------------


def test_every_asset_is_unknown_not_up(inventory):
    """The export says what exists, never what is running.

    Marking these `up` would inflate every availability figure on the site with
    machines nothing is actually watching.
    """
    assert {h["status"] for h in inventory.hosts} == {HostStatus.unknown}


def test_huawei_is_not_a_live_platform():
    from app.models import LIVE_PLATFORMS, PLATFORM_ORDER

    assert "huawei" in PLATFORM_ORDER
    assert "huawei" not in LIVE_PLATFORMS


# --- Credentials ------------------------------------------------------------


def test_no_credential_ever_reaches_a_host_record(inventory):
    """The workbook carries passwords and account names. None may be stored."""
    import json

    blob = json.dumps(inventory.hosts, default=str).lower()
    for secret in ("supersecret", "hunter2", "password", "account name"):
        assert secret not in blob, f"{secret!r} leaked into the imported records"


def test_user_columns_are_refused_by_name():
    """Even asked for directly, credential columns return nothing."""
    from app.huawei_assets import _pick

    headers = ["*Name", "User", "Password", "Security User"]
    row = ("host-1", "Administrator", "hunter2", "snmpUser")
    assert _pick(headers, row, "*Name") == "host-1"
    assert _pick(headers, row, "Password") == ""
    assert _pick(headers, row, "User") == ""
    assert _pick(headers, row, "Security User") == ""


# --- Robustness -------------------------------------------------------------


def test_a_missing_file_is_reported_not_raised(tmp_path):
    result = parse_workbook(tmp_path / "nope.xlsx")
    assert result.count == 0 and result.exported_at is None


def test_workbook_without_the_optional_sheets_still_parses(tmp_path):
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("PM Operating System")
    for _ in range(5):
        ws.append([])
    ws.append(["*Name", "IP", "Site name", "CPU", "Memory(MB)"])
    ws.append(["", "", "SiteName", "CPU", "Memory"])
    ws.append(["lonely-01", "10.9.5.5", "Site-PR", 4, 8192])
    path = tmp_path / "minimal.xlsx"
    wb.save(path)

    result = parse_workbook(path)
    assert result.count == 1
    host = result.hosts[0]
    assert host["hostname"] == "lonely-01"
    assert host["metrics"] == {"cores": 4, "mem_total_gb": 8.0}  # no disk sheet


# --- Into the database ------------------------------------------------------


def test_assets_land_as_huawei_hosts(client, tmp_path_factory):
    from sqlalchemy import select

    from app.db import SessionLocal
    from app.huawei_assets import load_into_db
    from app.models import Host, SourcePlatform

    path = _template(tmp_path_factory.mktemp("huawei-db"))
    db = SessionLocal()
    try:
        result = load_into_db(db, "Huawei-TEST", path)
        db.commit()
        rows = db.scalars(
            select(Host).where(Host.source_instance == "Huawei-TEST")
        ).all()
    finally:
        db.close()

    assert result.count == 4 and len(rows) == 4
    assert {r.source_platform for r in rows} == {SourcePlatform.huawei}
    assert {r.status for r in rows} == {HostStatus.unknown}
    assert next(r for r in rows if r.hostname == "vm-app-01").ip == "10.9.0.1"


def test_reimporting_the_same_export_does_not_duplicate(client, tmp_path_factory):
    from sqlalchemy import func, select

    from app.db import SessionLocal
    from app.huawei_assets import load_into_db
    from app.models import Host

    path = _template(tmp_path_factory.mktemp("huawei-idem"))
    db = SessionLocal()
    try:
        for _ in range(3):
            load_into_db(db, "Huawei-IDEM", path)
            db.commit()
        count = db.scalar(
            select(func.count(Host.id)).where(Host.source_instance == "Huawei-IDEM")
        )
    finally:
        db.close()
    assert count == 4
