"""Date-range filter (SQL) + branded xlsx export."""

from __future__ import annotations

import io
from datetime import datetime, timedelta, timezone

import openpyxl

from app.db import SessionLocal
from app.export_xlsx import build_workbook
from app.models import Alert, SourcePlatform
from app.routers.pages import _active_alerts, parse_dt

NOW = datetime(2026, 8, 12, 12, 0, tzinfo=timezone.utc)


def test_parse_dt_variants():
    assert parse_dt("2026-08-12T09:30") == datetime(2026, 8, 12, 9, 30, tzinfo=timezone.utc)
    assert parse_dt("2026-08-12") == datetime(2026, 8, 12, 0, 0, tzinfo=timezone.utc)
    assert parse_dt("") is None and parse_dt(None) is None and parse_dt("junk") is None


def test_build_workbook_structure_and_severity_color():
    data = build_workbook(
        sheet_title="Alerts",
        period="a → b",
        filters_summary="state=active",
        columns=["Severity", "Title"],
        rows=[["Disaster", "db down", 5], ["Warning", "cert", 2]],
        severity_col=0,
    )
    ws = openpyxl.load_workbook(io.BytesIO(data)).active
    assert ws["A1"].value == "SAMI'X — Monitoring Data Export"
    assert ws["A2"].value == "Extracted from SAMI'X tool"
    assert ws["A3"].value.startswith("Period:")
    assert ws.cell(6, 1).value == "Severity"       # header row
    assert ws.freeze_panes == "A7"
    assert ws.cell(7, 1).value == "Disaster"        # severity int stripped, not shown
    assert ws.cell(7, 2).value == "db down"
    assert ws.cell(7, 1).fill.fgColor.rgb.endswith("F6D0D0")  # critical tint


def test_alerts_date_range_filters_in_sql(client):
    inst = "SIS-DATE-TEST"
    db = SessionLocal()
    try:
        db.add(Alert(external_id="d-new", source_platform=SourcePlatform.zabbix,
                     source_instance=inst, severity_int=5, severity_label="Disaster",
                     title="recent", started_at=NOW, resolved=False))
        db.add(Alert(external_id="d-old", source_platform=SourcePlatform.zabbix,
                     source_instance=inst, severity_int=3, severity_label="Average",
                     title="stale", started_at=NOW - timedelta(hours=48), resolved=False))
        db.commit()
    finally:
        db.close()

    db = SessionLocal()
    try:
        # last 24h window -> only the recent one for this instance
        rows, *_ = _active_alerts(
            db, inst, 1, NOW - timedelta(hours=24), NOW + timedelta(hours=1)
        )
        titles = {a.title for a in rows}
        assert "recent" in titles and "stale" not in titles
    finally:
        db.close()


def test_alerts_state_filter_active_resolved_all(client):
    inst = "SIS-STATE-TEST"
    db = SessionLocal()
    try:
        db.add(Alert(external_id="s-open", source_platform=SourcePlatform.zabbix,
                     source_instance=inst, severity_int=5, severity_label="Disaster",
                     title="open one", started_at=NOW, resolved=False))
        db.add(Alert(external_id="s-done", source_platform=SourcePlatform.zabbix,
                     source_instance=inst, severity_int=3, severity_label="Average",
                     title="closed one", started_at=NOW, resolved=True))
        db.commit()
    finally:
        db.close()

    db = SessionLocal()
    try:
        active, *_ = _active_alerts(db, inst, 1, state="active")
        resolved, *_ = _active_alerts(db, inst, 1, state="resolved")
        every, *_ = _active_alerts(db, inst, 1, state="all")
        assert {a.title for a in active} == {"open one"}
        assert {a.title for a in resolved} == {"closed one"}
        assert {a.title for a in every} == {"open one", "closed one"}
    finally:
        db.close()


def test_dynatrace_export_dedups_zabbix_and_has_disk_column(client):
    from app.models import Host, HostStatus

    db = SessionLocal()
    try:
        for eid, pf, name, ip in [
            ("dz-z1", SourcePlatform.zabbix, "dz-shared", "10.99.0.1"),
            ("dz-d1", SourcePlatform.dynatrace, "dz-shared-dt", "10.99.0.1"),
            ("dz-d2", SourcePlatform.dynatrace, "dz-only-dt", "10.99.0.2"),
        ]:
            db.add(Host(external_id=eid, source_platform=pf, source_instance="DZT",
                        hostname=name, ip=ip, status=HostStatus.up))
        db.commit()
    finally:
        db.close()

    def sheet(url):
        r = client.get(url)
        assert r.status_code == 200
        ws = openpyxl.load_workbook(io.BytesIO(r.content)).active
        hdr = [ws.cell(6, i + 1).value for i in range(ws.max_column)]
        names = {ws.cell(r, 1).value for r in range(7, ws.max_row + 1)
                 if ws.cell(r, 1).value}
        return hdr, names

    hdr, names = sheet("/api/v1/capacity.xlsx?platform=dynatrace&q=dz-")
    assert hdr[-1] == "Disks (used / total GB)"
    assert {"dz-shared-dt", "dz-only-dt"} <= names

    _, names = sheet("/api/v1/capacity.xlsx?platform=dynatrace&q=dz-&dedup_zabbix=true")
    assert "dz-only-dt" in names
    assert "dz-shared-dt" not in names  # same IP as the Zabbix host -> dropped


def test_xlsx_endpoints_stream_valid_workbooks(client):
    for url, sheet in [("/api/v1/capacity.xlsx", "Capacity"),
                       ("/api/v1/alerts.xlsx?active=true", "Alerts")]:
        r = client.get(url)
        assert r.status_code == 200
        assert r.headers["content-type"].startswith("application/vnd.openxml")
        assert 'filename="SAMIX_' in r.headers["content-disposition"]
        ws = openpyxl.load_workbook(io.BytesIO(r.content)).active
        assert ws["A1"].value == "SAMI'X — Monitoring Data Export"
        assert ws.title == sheet
