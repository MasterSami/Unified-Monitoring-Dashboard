"""Per-host Zabbix capacity detail service (ported weekly-report logic)."""

from __future__ import annotations

import io

import openpyxl

from app.db import SessionLocal
from app.models import Host, HostStatus, SourcePlatform
from app.zabbix_report import (
    GB,
    REPORT_COLUMNS,
    _cache,
    host_capacity_detail,
    key_params,
    weekly_report_rows,
)


def _item(itemid, key, name, val):
    return {"itemid": itemid, "hostid": "1", "key_": key, "name": name, "lastvalue": val}


class _FakeReportZbx:
    instance = "Zabbix-66"

    def _rpc(self, method, params):
        if method == "host.get":
            return [{
                "hostid": "1", "host": "srv1", "name": "srv1",
                "interfaces": [{"main": "1", "ip": "10.0.0.1"}],
                "hostgroups": [{"name": "Linux servers"}],
                "tags": [{"tag": "Owner", "value": "NetTeam"}], "inventory": {}, "macros": [],
            }]
        if method == "item.get":
            return [
                _item("n", "system.cpu.num", "cores", "8"),
                _item("cpu", "system.cpu.util", "cpu", "17"),
                _item("mt", "vm.memory.size[total]", "mt", str(16 * GB)),
                _item("mu", "vm.memory.utilization", "mu", "75"),
                _item("ft", "vfs.fs.size[/,total]", "/ : Total space", str(100 * GB)),
                _item("fu", "vfs.fs.size[/,used]", "/ : Used space", str(40 * GB)),
                _item("vt", "vfs.fs.size[/var,total]", "/var : Total space", str(50 * GB)),
                _item("vu", "vfs.fs.size[/var,used]", "/var : Used space", str(10 * GB)),
            ]
        if method == "trend.get":
            return [
                {"itemid": "cpu", "num": "10", "value_min": "5", "value_avg": "17", "value_max": "40"},
                {"itemid": "mu", "num": "10", "value_min": "60", "value_avg": "75", "value_max": "90"},
            ]
        return []


def test_weekly_report_rows_batch():
    _cache.clear()
    rows = weekly_report_rows(_FakeReportZbx(), 7)
    assert len(rows) == 2 and len(rows[0]) == len(REPORT_COLUMNS)
    r = dict(zip(REPORT_COLUMNS, rows[0]))
    assert r["Total Cpu Cores"] == 8 and r["Memory Usage Percentage"] == 75
    assert r["Total VM Disk Size"] == 150.0 and r["Total VM Used Disk Size"] == 50.0
    assert r["Owner"] == "NetTeam" and "Linux servers" in r["Service"]
    assert r["Max Cpu Usage"] == "40.0%" and r["Avg Cpu Usage"] == "17%"
    assert r["VM Space For Each Drive"].startswith("/ : 100.0")


def test_capacity_report_endpoint_dynatrace_besteffort(client):
    db = SessionLocal()
    try:
        db.add(Host(hostname="dtsrv", ip="10.9.9.9", source_platform=SourcePlatform.dynatrace,
                    source_instance="Dynatrace-IT", external_id="dt1", status=HostStatus.up,
                    cpu_pct=33.0, mem_pct=64.0, metrics={"cores": 4, "mem_total_gb": 8.0}))
        db.commit()
    finally:
        db.close()
    r = client.get("/api/v1/capacity_report.xlsx?platform=dynatrace")
    assert r.status_code == 200
    ws = openpyxl.load_workbook(io.BytesIO(r.content)).active
    assert ws.title == "Capacity Report"
    assert [c.value for c in ws[6]] == REPORT_COLUMNS
    body = [row for row in ws.iter_rows(min_row=7, values_only=True) if row[0] == "dtsrv"]
    assert body and body[0][2] == 4  # cores from metrics


def test_key_params_quote_and_bracket_aware():
    assert key_params("vfs.fs.size[/var,total]") == ["/var", "total"]
    assert key_params('vfs.fs.size["C:",pused]') == ["C:", "pused"]
    assert key_params("system.cpu.num") == []


class _FakeZbx:
    """Stands in for ZabbixCollector — records _rpc calls, returns crafted data."""

    def __init__(self):
        self.instance = "Zabbix-66"
        self.calls = 0

    def _rpc(self, method, params):
        self.calls += 1
        if method == "host.get":
            return [{"name": "srv1", "interfaces": [{"main": "1", "ip": "10.0.0.1"}]}]
        if method == "item.get":
            return [
                {"itemid": "n", "key_": "system.cpu.num", "name": "CPU cores", "lastvalue": "8"},
                {"itemid": "cpu", "key_": "system.cpu.util", "name": "CPU util", "lastvalue": "17"},
                {"itemid": "mt", "key_": "vm.memory.size[total]", "name": "Total memory", "lastvalue": str(16 * GB)},
                {"itemid": "mu", "key_": "vm.memory.utilization", "name": "Memory util", "lastvalue": "75"},
                {"itemid": "ft", "key_": "vfs.fs.size[/,total]", "name": "/ : Total space", "lastvalue": str(100 * GB)},
                {"itemid": "fu", "key_": "vfs.fs.size[/,used]", "name": "/ : Used space", "lastvalue": str(40 * GB)},
            ]
        if method == "trend.get":
            return [
                {"itemid": "cpu", "num": "10", "value_min": "5", "value_avg": "17", "value_max": "40"},
                {"itemid": "mu", "num": "10", "value_min": "60", "value_avg": "75", "value_max": "90"},
            ]
        return []


def test_host_capacity_detail_parsing():
    _cache.clear()
    z = _FakeZbx()
    d = host_capacity_detail(z, "10001")
    assert d["host"] == "srv1" and d["ip"] == "10.0.0.1"
    assert d["cores"] == 8 and d["cpu_pct"] == 17.0
    assert d["mem_total_gb"] == 16.0 and d["mem_pct"] == 75.0
    assert d["disk_total_gb"] == 100.0 and d["disk_used_gb"] == 40.0 and d["disk_pct"] == 40.0
    assert d["cpu_trend"] == {"min": 5.0, "avg": 17.0, "max": 40.0}
    assert d["mem_trend"] == {"min": 60.0, "avg": 75.0, "max": 90.0}
    assert d["filesystems"] == [{"label": "/", "total_gb": 100.0, "used_gb": 40.0, "pct": 40.0}]


def test_detail_is_cached_per_host():
    _cache.clear()
    z = _FakeZbx()
    host_capacity_detail(z, "10001")
    first = z.calls
    host_capacity_detail(z, "10001")   # served from cache -> no more _rpc calls
    assert z.calls == first


def test_pavailable_memory_is_inverted():
    _cache.clear()

    class Z(_FakeZbx):
        def _rpc(self, method, params):
            if method == "item.get":
                return [
                    {"itemid": "ma", "key_": "vm.memory.size[pavailable]", "name": "avail", "lastvalue": "30"},
                ]
            if method == "trend.get":
                return [{"itemid": "ma", "num": "5", "value_min": "20", "value_avg": "30", "value_max": "40"}]
            return super()._rpc(method, params)

    d = host_capacity_detail(Z(), "1")
    assert d["mem_pct"] == 70.0                    # 100 - 30 available
    assert d["mem_trend"] == {"min": 60.0, "avg": 70.0, "max": 80.0}  # inverted
