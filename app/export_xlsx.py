"""Branded .xlsx export for SAMI'X (Capacity / Alerts).

One reusable builder writes a header block, then a styled data table (bold filled
header, freeze panes, readable dates, severity-colored cells) and returns the
workbook bytes for streaming. Rows are consumed from an iterator so a large
filtered result set is never materialized as a second Python list.
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_TITLE = "SAMI'X — Monitoring Data Export"
_HEADER_FILL = "1F4E78"          # table header band
_TITLE_FONT = Font(bold=True, size=16, color="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_MUTED = Font(color="666666", size=10)
_CREDIT_FONT = Font(color="1F4E78", size=10, bold=True, italic=True)

# Severity fills — aligned with the UI severity palette (5=critical … 1=ok).
_SEV_FILL = {
    5: "F6D0D0",  # critical
    4: "F7E0C8",  # major/high
    3: "F6EFC8",  # minor/average
    2: "D6E4F5",  # warning
    1: "D6F0DE",  # ok/info
}
_DATE_FMT = "yyyy-mm-dd hh:mm"


def build_workbook(
    *,
    sheet_title: str,
    period: str,
    filters_summary: str,
    columns: Sequence[str],
    rows: Iterable[Sequence[object]],
    severity_col: int | None = None,
    credit: str = "",
) -> bytes:
    """Return xlsx bytes: branded header block + styled table.

    ``severity_col`` (0-based index into ``columns``) tints each data row's
    severity cell by an accompanying integer the caller appends as the LAST
    value of every row (kept out of ``columns``); pass None for no coloring.

    ``credit`` adds an attribution line to the header block — Runbook exports
    use it to name the engineer who wrote the script.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    ncols = len(columns)
    last_col = get_column_letter(ncols)

    # --- header block ----------------------------------------------------
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = _TITLE
    c.font = _TITLE_FONT
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"].value = "Extracted from SAMI'X tool"
    ws["A2"].font = _MUTED

    ws.merge_cells(f"A3:{last_col}3")
    ws["A3"].value = f"Period: {period}"
    ws["A3"].font = _MUTED

    ws.merge_cells(f"A4:{last_col}4")
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws["A4"].value = f"Generated at: {now} | Filters: {filters_summary}"
    ws["A4"].font = _MUTED

    header_row = 6
    if credit:
        ws.merge_cells(f"A5:{last_col}5")
        ws["A5"].value = credit
        ws["A5"].font = _CREDIT_FONT
        header_row = 7
    fill = PatternFill("solid", start_color=_HEADER_FILL)
    for i, name in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=i, value=name)
        cell.font = _HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- data rows -------------------------------------------------------
    widths = [len(str(x)) for x in columns]
    r = header_row
    for row in rows:
        r += 1
        sev = None
        values = list(row)
        if severity_col is not None and values:
            sev = values.pop()  # trailing severity int, not a visible column
        for i, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=i, value=val)
            if isinstance(val, datetime):
                cell.number_format = _DATE_FMT
            if i - 1 < len(widths):
                widths[i - 1] = max(widths[i - 1], len(str(val)) if val is not None else 0)
        if sev is not None and severity_col is not None:
            fillcode = _SEV_FILL.get(int(sev))
            if fillcode:
                ws.cell(row=r, column=severity_col + 1).fill = PatternFill(
                    "solid", start_color=fillcode
                )

    # --- finishing touches ----------------------------------------------
    ws.freeze_panes = f"A{header_row + 1}"
    if r > header_row:
        ws.auto_filter.ref = f"A{header_row}:{last_col}{r}"
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(48, max(10, w + 2))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
