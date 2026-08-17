"""JSON API under ``/api/v1``.

These endpoints back the UI and are the integration surface for later reports
and automation.
"""

from __future__ import annotations

import csv
import io
import secrets
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, Header, HTTPException, Query
from fastapi.responses import Response, StreamingResponse
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.config import Settings, get_settings
from app.db import get_db
from app.models import (
    Alert,
    CollectorRun,
    Host,
    HostStatus,
    RunStatus,
    SourcePlatform,
    TopologyEdge,
    TopologyNode,
)
from app.normalizer import severity_label
from app.scheduler import get_collector_statuses, get_service, run_topology_now
from app.sitescope_ingest import ingest_lines
from app.topology import (
    NNMI_L2_COLUMNS,
    UNIFIED_COLUMNS,
    dynatrace_app_rows,
    dynatrace_service_rows,
    dynatrace_unified_rows,
    nnmi_connection_rows,
)
from app.schemas import (
    AlertOut,
    CollectorStatus,
    HostOut,
    IngestResult,
    PlatformHostCount,
    SeverityBucket,
    SiteScopeIngest,
    SummaryOut,
)

router = APIRouter(prefix="/api/v1", tags=["api"])


# --- SiteScope push ingest --------------------------------------------------


def _check_ingest_auth(authorization: str | None, settings: Settings) -> None:
    """Bearer-token auth for the ingest endpoint (constant-time compare)."""
    token = settings.sitescope_ingest_token
    if not token:
        raise HTTPException(status_code=503, detail="SiteScope ingest is not configured")
    expected = f"Bearer {token}"
    if not authorization or not secrets.compare_digest(authorization, expected):
        raise HTTPException(status_code=401, detail="invalid or missing bearer token")


@router.post("/ingest/sitescope", response_model=IngestResult)
def ingest_sitescope(
    payload: SiteScopeIngest,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
    settings: Settings = Depends(get_settings),
) -> IngestResult:
    """Receive a batch of redacted SiteScope log lines from the forwarder.

    Bearer-authenticated, size-capped, and idempotent: re-sending the same batch
    updates rows in place (no duplicates). A heartbeat (empty ``lines``) still
    records a collector run so a dead forwarder shows up as stale in the UI.
    """
    _check_ingest_auth(authorization, settings)
    if len(payload.lines) > settings.ingest_max_events:
        raise HTTPException(
            status_code=413,
            detail=f"too many events (>{settings.ingest_max_events})",
        )
    if sum(len(line) for line in payload.lines) > settings.ingest_max_bytes:
        raise HTTPException(status_code=413, detail="payload too large")

    started = datetime.now(timezone.utc)
    c = ingest_lines(db, payload.source_instance, payload.lines)

    # Collector-health heartbeat: one run row per ingest, so the dashboard can
    # tell "no alerts" (recent run, 0 events) from "collector dead" (stale run).
    db.add(
        CollectorRun(
            platform="sitescope",
            instance=payload.source_instance,
            started_at=started,
            finished_at=datetime.now(timezone.utc),
            status=RunStatus.success,
            items_collected=c.events,
            # Distinct hosts seen this batch (not just new inserts) so the
            # collector row shows the real fleet size, not 0 on repeat runs.
            hosts_collected=c.hosts,
            alerts_collected=c.events,
        )
    )
    db.commit()
    return IngestResult(
        status="ok",
        received=c.received,
        inserted=c.inserted,
        updated=c.updated,
        skipped=c.skipped,
        redactions=c.redactions,
    )


@router.get("/hosts", response_model=list[HostOut])
def list_hosts(
    platform: str | None = Query(default=None),
    status: str | None = Query(default=None),
    q: str | None = Query(default=None),
    db: Session = Depends(get_db),
) -> list[Host]:
    """Return hosts, optionally filtered by platform, status, or search text."""
    stmt = select(Host)
    if platform:
        stmt = stmt.where(Host.source_platform == platform)
    if status:
        stmt = stmt.where(Host.status == status)
    if q:
        like = f"%{q.lower()}%"
        stmt = stmt.where(
            func.lower(Host.hostname).like(like) | func.lower(Host.ip).like(like)
        )
    stmt = stmt.order_by(Host.hostname.asc())
    return list(db.scalars(stmt).all())


@router.get("/alerts", response_model=list[AlertOut])
def list_alerts(
    active: bool = Query(default=True),
    db: Session = Depends(get_db),
) -> list[Alert]:
    """Return alerts. ``active=true`` (default) excludes resolved alerts."""
    stmt = select(Alert)
    if active:
        stmt = stmt.where(Alert.resolved.is_(False))
    stmt = stmt.order_by(
        Alert.severity_int.desc(), Alert.started_at.desc().nullslast()
    )
    return list(db.scalars(stmt).all())


def _csv_response(filename: str, header: list[str], rows: list[list]) -> StreamingResponse:
    """Serialize rows to a downloadable CSV StreamingResponse."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(header)
    writer.writerows(rows)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _require_export(settings: Settings) -> None:
    """Reject the request when CSV export is disabled by config."""
    if not settings.enable_export:
        raise HTTPException(status_code=404, detail="CSV export is disabled")


@router.get("/hosts.csv")
def export_hosts_csv(
    platform: str | None = Query(default=None),
    instance: str | None = Query(default=None),
    status: str | None = Query(default=None),
    q: str | None = Query(default=None),
    db: Session = Depends(get_db),
    settings: Settings = Depends(get_settings),
) -> StreamingResponse:
    """Export hosts (respecting filters) as CSV."""
    _require_export(settings)
    stmt = select(Host)
    if platform and platform != "all":
        stmt = stmt.where(Host.source_platform == platform)
    if instance and instance != "all":
        stmt = stmt.where(Host.source_instance == instance)
    if status and status != "all":
        stmt = stmt.where(Host.status == status)
    if q:
        like = f"%{q.lower()}%"
        stmt = stmt.where(
            func.lower(Host.hostname).like(like)
            | func.lower(func.coalesce(Host.ip, "")).like(like)
            | func.lower(func.coalesce(Host.group_name, "")).like(like)
            | func.lower(func.coalesce(Host.source_instance, "")).like(like)
        )
    stmt = stmt.order_by(Host.hostname.asc())
    rows = [
        [
            h.hostname,
            h.ip or "",
            h.source_platform.value,
            h.source_instance,
            h.status.value,
            h.group_name or "",
            h.last_seen.isoformat() if h.last_seen else "",
        ]
        for h in db.scalars(stmt).all()
    ]
    return _csv_response(
        "hosts.csv",
        ["hostname", "ip", "platform", "instance", "status", "group", "last_seen"],
        rows,
    )


@router.get("/capacity.csv")
def export_capacity_csv(
    platform: str | None = Query(default=None),
    instance: str | None = Query(default=None),
    group: str | None = Query(default=None),
    status: str | None = Query(default=None),
    q: str | None = Query(default=None),
    db: Session = Depends(get_db),
    settings: Settings = Depends(get_settings),
) -> StreamingResponse:
    """Export the capacity view (hosts + CPU/mem/disk) as CSV."""
    _require_export(settings)
    stmt = select(Host)
    if platform and platform != "all":
        stmt = stmt.where(Host.source_platform == platform)
    if instance and instance != "all":
        stmt = stmt.where(Host.source_instance == instance)
    if group and group != "all":
        stmt = stmt.where(Host.group_name == group)
    if status and status != "all":
        stmt = stmt.where(Host.status == status)
    if q:
        like = f"%{q.lower()}%"
        stmt = stmt.where(
            func.lower(Host.hostname).like(like)
            | func.lower(func.coalesce(Host.ip, "")).like(like)
            | func.lower(func.coalesce(Host.group_name, "")).like(like)
            | func.lower(func.coalesce(Host.source_instance, "")).like(like)
        )
    stmt = stmt.order_by(Host.hostname.asc())
    rows = [
        [
            h.hostname,
            h.ip or "",
            h.source_platform.value,
            h.source_instance,
            h.group_name or "",
            "" if h.cpu_pct is None else h.cpu_pct,
            "" if h.mem_pct is None else h.mem_pct,
            "" if h.disk_pct is None else h.disk_pct,
            (h.metrics or {}).get("cores", ""),
            (h.metrics or {}).get("mem_total_gb", ""),
            h.status.value,
        ]
        for h in db.scalars(stmt).all()
    ]
    return _csv_response(
        "capacity.csv",
        [
            "server",
            "ip",
            "platform",
            "instance",
            "group",
            "cpu_pct",
            "mem_pct",
            "disk_pct",
            "cores",
            "mem_total_gb",
            "status",
        ],
        rows,
    )


@router.get("/alerts.csv")
def export_alerts_csv(
    active: bool = Query(default=True),
    q: str | None = Query(default=None),
    db: Session = Depends(get_db),
    settings: Settings = Depends(get_settings),
) -> StreamingResponse:
    """Export alerts (respecting filters) as CSV."""
    _require_export(settings)
    stmt = select(Alert)
    if active:
        stmt = stmt.where(Alert.resolved.is_(False))
    if q:
        like = f"%{q.lower()}%"
        stmt = stmt.where(
            func.lower(Alert.title).like(like)
            | func.lower(func.coalesce(Alert.host_hostname, "")).like(like)
            | func.lower(func.coalesce(Alert.source_instance, "")).like(like)
            | func.lower(Alert.source_platform).like(like)
            | func.lower(Alert.severity_label).like(like)
        )
    stmt = stmt.order_by(
        Alert.severity_int.desc(), Alert.started_at.desc().nullslast()
    )
    rows = [
        [
            a.severity_int,
            a.severity_label,
            a.title,
            a.source_platform.value,
            a.source_instance,
            a.host_hostname or "",
            a.started_at.isoformat() if a.started_at else "",
            "resolved" if a.resolved else "active",
        ]
        for a in db.scalars(stmt).all()
    ]
    return _csv_response(
        "alerts.csv",
        [
            "severity_int",
            "severity_label",
            "title",
            "platform",
            "instance",
            "host",
            "started_at",
            "state",
        ],
        rows,
    )


# --- Branded Excel export ---------------------------------------------------

_XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _fname_stamp(value: str | None) -> str:
    """A filesystem-safe compact stamp from a datetime-local string (or 'all')."""
    if not value:
        return "all"
    return "".join(c for c in value if c.isalnum())


def _period_str(date_from: str | None, date_to: str | None) -> str:
    return f"{date_from or 'earliest'} → {date_to or 'now'}"


def _xlsx_response(name: str, data: bytes) -> Response:
    return Response(
        content=data,
        media_type=_XLSX_MEDIA,
        headers={"Content-Disposition": f'attachment; filename="{name}"'},
    )


@router.get("/capacity.xlsx")
def export_capacity_xlsx(
    platform: str | None = Query(default=None),
    instance: str | None = Query(default=None),
    group: str | None = Query(default=None),
    status: str | None = Query(default=None),
    q: str | None = Query(default=None),
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    db: Session = Depends(get_db),
    settings: Settings = Depends(get_settings),
) -> Response:
    """Export the filtered Capacity view as a branded .xlsx (same SQL filters)."""
    _require_export(settings)
    from app.export_xlsx import build_workbook
    from app.routers.pages import _hosts_stmt, parse_dt

    stmt = _hosts_stmt(q, platform, status, instance, group,
                       parse_dt(date_from), parse_dt(date_to))
    stmt = stmt.order_by(Host.hostname.asc())

    def rows():
        for h in db.scalars(stmt):
            m = h.metrics or {}
            yield [
                h.hostname, h.ip or "", h.source_platform.value, h.source_instance,
                h.group_name or "",
                "" if h.cpu_pct is None else h.cpu_pct,
                "" if h.mem_pct is None else h.mem_pct,
                "" if h.disk_pct is None else h.disk_pct,
                m.get("cores", ""), m.get("mem_total_gb", ""),
                h.last_seen, h.status.value,
            ]

    filters = ", ".join(
        f"{k}={v}" for k, v in (
            ("platform", platform), ("instance", instance), ("group", group),
            ("status", status), ("q", q),
        ) if v and v != "all"
    ) or "none"
    data = build_workbook(
        sheet_title="Capacity",
        period=_period_str(date_from, date_to),
        filters_summary=filters,
        columns=["Server", "IP", "Platform", "Instance", "Group", "CPU %",
                 "Memory %", "Disk %", "Cores", "Total Memory (GB)",
                 "Last Updated", "Status"],
        rows=rows(),
    )
    fname = f"SAMIX_capacity_{_fname_stamp(date_from)}_{_fname_stamp(date_to)}.xlsx"
    return _xlsx_response(fname, data)


@router.get("/alerts.xlsx")
def export_alerts_xlsx(
    state: str = Query(default="active"),
    active: bool | None = Query(default=None),
    q: str | None = Query(default=None),
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    db: Session = Depends(get_db),
    settings: Settings = Depends(get_settings),
) -> Response:
    """Export the filtered Alerts view as a branded, severity-colored .xlsx.

    ``state`` is active/resolved/all. ``active`` is kept for backward compat:
    ``active=false`` maps to ``all`` when ``state`` isn't given explicitly.
    """
    _require_export(settings)
    from app.export_xlsx import build_workbook
    from app.routers.pages import _alerts_filtered_stmt, parse_dt

    if active is False and state == "active":
        state = "all"
    if state not in ("active", "resolved", "all"):
        state = "active"

    stmt = _alerts_filtered_stmt(q, parse_dt(date_from), parse_dt(date_to), state)
    stmt = stmt.order_by(
        Alert.severity_int.desc(), Alert.started_at.desc().nullslast()
    )

    def rows():
        for a in db.scalars(stmt):
            # trailing severity_int drives the severity-cell color (not a column)
            yield [
                a.severity_label, a.title, a.source_platform.value,
                a.source_instance or "", a.host_hostname or "", a.started_at,
                "resolved" if a.resolved else "active", a.severity_int,
            ]

    filters = ", ".join(
        f"{k}={v}" for k, v in (("state", state), ("q", q)) if v
    ) or "none"
    data = build_workbook(
        sheet_title="Alerts",
        period=_period_str(date_from, date_to),
        filters_summary=filters,
        columns=["Severity", "Title", "Platform", "Instance", "Host",
                 "Started", "State"],
        rows=rows(),
        severity_col=0,
    )
    fname = f"SAMIX_alerts_{_fname_stamp(date_from)}_{_fname_stamp(date_to)}.xlsx"
    return _xlsx_response(fname, data)


def _db_report_rows(db: Session, platform: str) -> list[list]:
    """Best-effort weekly-report rows from stored host records (one per host).

    Used for Dynatrace (and as a fallback) where per-filesystem/trend data isn't
    pulled live: fills the same columns from the Host snapshot, N/A where unknown.
    """
    from app.zabbix_report import REPORT_COLUMNS  # noqa: F401 (column order)

    stmt = select(Host).where(Host.source_platform == platform).order_by(Host.hostname.asc())
    rows: list[list] = []
    for h in db.scalars(stmt):
        m = h.metrics or {}
        rows.append([
            h.hostname, h.ip or "N/A",
            m.get("cores", 0), h.cpu_pct if h.cpu_pct is not None else 0,
            m.get("mem_total_gb", 0.0), int(h.mem_pct) if h.mem_pct is not None else 0,
            m.get("disk_total_gb", 0.0), m.get("disk_used_gb", 0.0),
            "N/A", "N/A",
            h.group_name or "N/A", "N/A",
            "N/A", "N/A", "N/A", "N/A", "N/A", "N/A",
        ])
    return rows


def _dynatrace_report_rows(db: Session, settings: Settings) -> list[list]:
    """Weekly-report rows for Dynatrace hosts, one row per partition.

    Host-level fields come from the stored snapshot; per-partition disk usage
    (C:, D:, /…) is pulled live from any configured Dynatrace collector so the
    "VM Space / VM Used Space For Each Drive" columns match the Zabbix report.
    Best-effort: if the live breakdown is unavailable, each host gets a single
    row with the summed disk total.
    """
    from app.servers import load_servers

    breakdown: dict[str, list] = {}
    if not settings.mock_mode:
        for s in load_servers(settings):
            if getattr(s, "platform", "") != "dynatrace":
                continue
            collector = get_service().get(s.name)
            if collector is None or getattr(collector, "name", "") != "dynatrace":
                continue
            try:
                for hid, disks in collector.disk_breakdown().items():
                    breakdown.setdefault(hid, []).extend(disks)
            except Exception:  # noqa: BLE001 — one instance failing is non-fatal
                continue

    rows: list[list] = []
    stmt = (
        select(Host)
        .where(Host.source_platform == "dynatrace")
        .order_by(Host.hostname.asc())
    )
    for h in db.scalars(stmt):
        m = h.metrics or {}
        base = [
            h.hostname, h.ip or "N/A",
            m.get("cores", 0), h.cpu_pct if h.cpu_pct is not None else 0,
            m.get("mem_total_gb", 0.0), int(h.mem_pct) if h.mem_pct is not None else 0,
            m.get("disk_total_gb", 0.0), m.get("disk_used_gb", 0.0),
            "N/A", "N/A",                       # per-drive (cols 8, 9) — filled below
            h.group_name or "N/A", "N/A",
            "N/A", "N/A", "N/A", "N/A", "N/A", "N/A",
        ]
        disks = breakdown.get(h.external_id)
        if disks:
            for label, total_gb, used_gb in sorted(disks):
                r = list(base)
                r[8] = f"{label} : {total_gb}"
                r[9] = f"{label} : {used_gb}"
                rows.append(r)
        else:
            rows.append(list(base))
    return rows


@router.get("/capacity_report.xlsx")
def export_capacity_report_xlsx(
    platform: str | None = Query(default="all"),
    instance: str | None = Query(default="all"),
    days: int = Query(default=7),
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    db: Session = Depends(get_db),
    settings: Settings = Depends(get_settings),
) -> Response:
    """Rich capacity export (the weekly-report columns).

    Zabbix rows are pulled LIVE per instance (cores, cpu/mem/disk, per-filesystem,
    and min/avg/max CPU/RAM trends over the window); Dynatrace is best-effort from
    the stored snapshot. Export only — the on-screen view is unchanged.
    """
    _require_export(settings)
    from app.export_xlsx import build_workbook
    from app.routers.pages import parse_dt
    from app.servers import load_servers
    from app.zabbix_report import REPORT_COLUMNS, weekly_report_rows

    df, dt = parse_dt(date_from), parse_dt(date_to)
    if df and dt and dt > df:
        days = max(1, (dt - df).days or 1)

    rows: list[list] = []

    # --- Zabbix: live report per selected instance ---
    zbx_instances: list[str] = []
    if not settings.mock_mode and platform in (None, "all", "zabbix"):
        if instance and instance != "all":
            zbx_instances = [instance]
        else:
            zbx_instances = [s.name for s in load_servers(settings) if s.platform == "zabbix"]
    for inst in zbx_instances:
        collector = get_service().get(inst)
        if collector is None or getattr(collector, "name", "") != "zabbix":
            continue
        try:
            rows.extend(weekly_report_rows(collector, days))
        except Exception:  # noqa: BLE001 — one instance failing must not kill the export
            continue

    # --- Dynatrace: snapshot host-level + live per-partition disk ---
    if platform in (None, "all", "dynatrace") and (instance in (None, "all") or platform == "dynatrace"):
        rows.extend(_dynatrace_report_rows(db, settings))

    # Mock/demo (no live Zabbix): fall back to the snapshot for the whole view.
    if settings.mock_mode and not rows:
        for pf in ("zabbix", "dynatrace"):
            rows.extend(_db_report_rows(db, pf))

    period = _period_str(date_from, date_to) if (df or dt) else f"trend window: last {days} days"
    filters = ", ".join(
        f"{k}={v}" for k, v in (("platform", platform), ("instance", instance)) if v and v != "all"
    ) or "none"
    data = build_workbook(
        sheet_title="Capacity Report",
        period=period,
        filters_summary=filters,
        columns=REPORT_COLUMNS,
        rows=rows,
    )
    fname = f"SAMIX_capacity_report_{_fname_stamp(date_from)}_{_fname_stamp(date_to)}.xlsx"
    return _xlsx_response(fname, data)


# --- Topology ---------------------------------------------------------------

#: Map the UI "view" name to the platform that owns that kind of graph.
_VIEW_PLATFORM = {
    "network": SourcePlatform.nnmi,
    "service": SourcePlatform.dynatrace,
}


def _require_topology(settings: Settings) -> None:
    """Reject the request when the topology feature is disabled by config."""
    if not settings.enable_topology:
        raise HTTPException(status_code=404, detail="Topology is disabled")


@router.get("/topology/graph")
def topology_graph(
    view: str = Query(default="network"),
    instance: str | None = Query(default=None),
    db: Session = Depends(get_db),
    settings: Settings = Depends(get_settings),
) -> dict:
    """Return the topology graph as Cytoscape.js elements JSON.

    ``view=network`` returns the NNMi L2 map; ``view=service`` returns the
    Dynatrace service dependency map. Filter to one instance with ``instance``.
    """
    _require_topology(settings)
    platform = _VIEW_PLATFORM.get(view, SourcePlatform.nnmi)

    node_stmt = select(TopologyNode).where(TopologyNode.source_platform == platform)
    edge_stmt = select(TopologyEdge).where(TopologyEdge.source_platform == platform)
    if instance and instance != "all":
        node_stmt = node_stmt.where(TopologyNode.source_instance == instance)
        edge_stmt = edge_stmt.where(TopologyEdge.source_instance == instance)

    nodes = list(db.scalars(node_stmt).all())
    edges = list(db.scalars(edge_stmt).all())
    # A node is keyed per instance, so scope the id by instance to stay unique
    # even when two instances share an external_id.
    def nid(inst: str, ext: str) -> str:
        return f"{inst}::{ext}"

    known = {nid(n.source_instance, n.external_id) for n in nodes}
    el_nodes = [
        {
            "data": {
                "id": nid(n.source_instance, n.external_id),
                "label": n.name,
                "kind": n.kind,
                "category": n.category or "",
                "status": (n.status or "").upper(),
                "instance": n.source_instance,
            }
        }
        for n in nodes
    ]
    el_edges = []
    for e in edges:
        src = nid(e.source_instance, e.from_external_id)
        dst = nid(e.source_instance, e.to_external_id)
        if src not in known or dst not in known:
            continue
        el_edges.append(
            {
                "data": {
                    "id": nid(e.source_instance, e.external_id),
                    "source": src,
                    "target": dst,
                    "label": e.label or "",
                    "kind": e.kind,
                }
            }
        )
    return {
        "view": view,
        "platform": platform.value,
        "instance": instance or "all",
        "counts": {"nodes": len(el_nodes), "edges": len(el_edges)},
        "elements": {"nodes": el_nodes, "edges": el_edges},
    }


@router.post("/topology/run")
def topology_run(settings: Settings = Depends(get_settings)) -> dict[str, str]:
    """Rebuild every instance's topology graph now (synchronous)."""
    _require_topology(settings)
    run_topology_now()
    return {"status": "ok"}


def _require_topology_export(settings: Settings) -> None:
    """Reject the request when topology export is disabled by config."""
    if not settings.enable_topology or not settings.enable_topology_export:
        raise HTTPException(status_code=404, detail="Topology export is disabled")


def _topology_edges(
    db: Session, platform: SourcePlatform, instance: str | None
) -> list[TopologyEdge]:
    stmt = select(TopologyEdge).where(TopologyEdge.source_platform == platform)
    if instance and instance != "all":
        stmt = stmt.where(TopologyEdge.source_instance == instance)
    return list(db.scalars(stmt).all())


@router.get("/topology/nnmi-l2.csv")
def export_nnmi_l2_csv(
    instance: str | None = Query(default=None),
    db: Session = Depends(get_db),
    settings: Settings = Depends(get_settings),
) -> StreamingResponse:
    """Export the NNMi L2 connections (name, status, endpoints, interfaces)."""
    _require_topology_export(settings)
    edges = _topology_edges(db, SourcePlatform.nnmi, instance)
    rows = nnmi_connection_rows(edges)
    return _csv_response(
        "nnmi_l2_connections.csv",
        NNMI_L2_COLUMNS,
        [[r[c] for c in NNMI_L2_COLUMNS] for r in rows],
    )


@router.get("/topology/dynatrace-map.xlsx")
def export_dynatrace_map_xlsx(
    instance: str | None = Query(default=None),
    db: Session = Depends(get_db),
    settings: Settings = Depends(get_settings),
) -> Response:
    """Export the Dynatrace unified map as an XLSX with three sheets.

    Sheets: ``Unified_Map`` (per-path), ``App_to_App`` and ``Service_to_Service``
    — the same shape as the ``Dynatrace_Inventory_Map.xlsx`` export.
    """
    _require_topology_export(settings)
    from openpyxl import Workbook  # local import: only needed for export
    from openpyxl.styles import Font, PatternFill

    edges = _topology_edges(db, SourcePlatform.dynatrace, instance)
    unified = dynatrace_unified_rows(edges)
    app2app = dynatrace_app_rows(unified)
    svc2svc = dynatrace_service_rows(unified)

    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF")

    def _sheet(wb, title, columns, rows):
        ws = wb.create_sheet(title)
        ws.append([label for _key, label in columns])
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill, cell.font = hdr_fill, hdr_font
        for r in rows:
            ws.append([r.get(key, "") for key, _label in columns])
        ws.freeze_panes = "A2"
        if ws.max_row > 1:
            ws.auto_filter.ref = ws.dimensions
        return ws

    wb = Workbook()
    wb.remove(wb.active)
    _sheet(wb, "Unified_Map", UNIFIED_COLUMNS, unified)
    _sheet(
        wb,
        "App_to_App",
        [
            ("source_application", "Source Application"),
            ("target_application", "Target Application"),
            ("link_type", "Link Type"),
        ],
        app2app,
    )
    _sheet(
        wb,
        "Service_to_Service",
        [
            ("source_service", "Source Service"),
            ("target_service", "Target Service"),
            ("link_type", "Link Type"),
            ("middleware_chain", "Middleware Chain"),
        ],
        svc2svc,
    )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="dynatrace_unified_map.xlsx"'
        },
    )


@router.get("/summary", response_model=SummaryOut)
def summary(
    db: Session = Depends(get_db),
    settings: Settings = Depends(get_settings),
) -> SummaryOut:
    """Return aggregate KPIs for the overview page."""
    total_hosts = db.scalar(select(func.count(Host.id))) or 0
    hosts_down = (
        db.scalar(
            select(func.count(Host.id)).where(Host.status == HostStatus.down)
        )
        or 0
    )
    active_alerts = (
        db.scalar(select(func.count(Alert.id)).where(Alert.resolved.is_(False)))
        or 0
    )

    # Per-platform host totals and down counts (two small grouped queries).
    def _by_platform(rows: list[tuple]) -> dict[str, int]:
        return {getattr(p, "value", str(p)): int(c) for p, c in rows}

    total_by_platform = _by_platform(
        db.execute(
            select(Host.source_platform, func.count(Host.id)).group_by(
                Host.source_platform
            )
        ).all()
    )
    down_by_platform = _by_platform(
        db.execute(
            select(Host.source_platform, func.count(Host.id))
            .where(Host.status == HostStatus.down)
            .group_by(Host.source_platform)
        ).all()
    )
    per_platform = [
        PlatformHostCount(
            platform=plat,
            total=total_by_platform.get(plat, 0),
            down=down_by_platform.get(plat, 0),
        )
        for plat in ("zabbix", "dynatrace", "nnmi", "sitescope")
    ]

    # Active alerts by severity.
    sev_rows = db.execute(
        select(Alert.severity_int, func.count(Alert.id))
        .where(Alert.resolved.is_(False))
        .group_by(Alert.severity_int)
    ).all()
    sev_counts = {int(s): int(c) for s, c in sev_rows}
    severity_buckets = [
        SeverityBucket(
            severity_int=level,
            label=severity_label(level),
            count=sev_counts.get(level, 0),
        )
        for level in (5, 4, 3, 2, 1)
    ]

    collectors = get_collector_statuses(db, settings)

    return SummaryOut(
        total_hosts=total_hosts,
        hosts_down=hosts_down,
        active_alerts=active_alerts,
        per_platform=per_platform,
        severity_buckets=severity_buckets,
        collectors=collectors,
    )


@router.get("/collectors/status", response_model=list[CollectorStatus])
def collectors_status(
    db: Session = Depends(get_db),
    settings: Settings = Depends(get_settings),
) -> list[CollectorStatus]:
    """Return current health for every configured instance."""
    return get_collector_statuses(db, settings)


@router.post("/collectors/{instance}/run")
def run_collector(instance: str) -> dict[str, str]:
    """Manually trigger a single instance's collection run (synchronous)."""
    service = get_service()
    ran = service.run_one(instance)
    if not ran:
        raise HTTPException(status_code=404, detail=f"unknown instance: {instance}")
    return {"status": "ok", "instance": instance}


@router.post("/collectors/run")
def run_all_collectors() -> dict[str, object]:
    """Trigger a run of every configured instance (synchronous)."""
    service = get_service()
    service.run_all()
    return {"status": "ok", "instances": list(service.collectors)}


@router.post("/collectors/{instance}/test-mail")
def test_mail(
    instance: str,
    to: str | None = Query(default=None),
    settings: Settings = Depends(get_settings),
) -> dict[str, object]:
    """Ask a Zabbix instance to send a test email (verifies email alerting)."""
    service = get_service()
    collector = service.get(instance)
    if collector is None or not hasattr(collector, "send_test_mail"):
        raise HTTPException(
            status_code=404,
            detail=f"instance '{instance}' not found or does not support test mail",
        )
    recipient = to or settings.test_mail_to
    if not recipient:
        raise HTTPException(
            status_code=400,
            detail="no recipient: pass ?to=... or set TEST_MAIL_TO in .env",
        )
    result = collector.send_test_mail(recipient)  # type: ignore[attr-defined]
    return {"instance": instance, "to": recipient, **result}
