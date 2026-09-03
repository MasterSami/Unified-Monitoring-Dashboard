"""Huawei i2000 / Digital View asset inventory, read from an exported workbook.

Huawei keeps the API port closed to us, so the inventory arrives the only way it
can: someone exports ``BaseAssetImportTemplate_En.xlsx`` from the Digital View
UI and drops it on disk. This module turns that workbook into the same host
shape every collector produces, so Huawei assets appear on the Hosts, Capacity
and Shared pages like anything else.

**This is inventory, not monitoring.** The export says what exists, how much CPU
and memory it has, and where it sits in the rack. It never says whether anything
is running right now. So every host here is stored with ``unknown`` status —
the same honest answer the dashboard gives for any host whose state it cannot
observe. Marking them ``up`` would inflate "Total Active Servers" with machines
nothing is actually watching, which is the exact mistake the Dynatrace
discovered-host count already taught us not to make.

Reading the workbook
--------------------
The file is a Huawei import template, so its sheets have five rows of metadata
before the data starts:

    row 1-2   type ids and group names
    row 3-5   section headers, spanning merged cells
    row 6     the real column headers
    row 7     a row of internal field ids ("SiteName", "HostName", …)
    row 8+    actual assets

The ``*Information`` sheets are shallower — header on row 3, field ids on row 4,
data from row 5. Both shapes are handled below.

Credentials
-----------
The workbook carries account columns (``Account name``, ``Password``, SNMP
``Authentication Password``…). **Nothing from those columns is read or stored.**
Only the asset fields listed in the extractors below are touched.
"""

from __future__ import annotations

import logging
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path

from app.models import HostStatus, SourcePlatform

logger = logging.getLogger("huawei")

#: Sheets holding host-like assets, and the row their real header sits on.
_HOST_SHEETS = {
    "VM Operating System": 6,
    "PM Operating System": 6,
}
#: Sheets holding physical equipment (servers in racks, storage arrays).
_DEVICE_SHEETS = {
    "Rack Server": 6,
    "Storage Device": 6,
    "Network Device": 6,
    "F5": 6,
    "SHELF": 6,
}
#: Detail sheets, keyed by asset name.
_DETAIL_HEADER_ROW = 3

#: Column headers that must never be read, whatever sheet they appear on.
_FORBIDDEN = {
    "password", "*password", "account name", "*account name",
    "authentication password", "encryption password", "oracle user password",
    "cluster grid password", "user", "security user", "switch account",
}


@dataclass
class HuaweiInventory:
    """What one workbook contained."""

    hosts: list[dict] = field(default_factory=list)
    exported_at: datetime | None = None
    sheets_read: dict[str, int] = field(default_factory=dict)
    skipped: int = 0

    @property
    def count(self) -> int:
        return len(self.hosts)


def _clean(value: object) -> str:
    """Normalize a cell to a trimmed string ('' for blanks)."""
    if value is None:
        return ""
    return str(value).strip()


def _rows(ws, header_row: int) -> tuple[list[str], list[tuple]]:
    """Return (headers, data rows) for a sheet, skipping the field-id row.

    Row ``header_row + 1`` holds Huawei's internal field ids rather than data,
    so it is dropped. It is recognised by having no value in the first column
    while later columns carry names like "SiteName" — the same shape a real
    asset never has, since every asset must have a name.
    """
    grid = list(ws.iter_rows(values_only=True))
    if len(grid) <= header_row:
        return [], []
    headers = [_clean(c) for c in grid[header_row - 1]]
    body: list[tuple] = []
    for row in grid[header_row:]:
        if not any(_clean(c) for c in row):
            continue
        if not _clean(row[0]):
            continue  # field-id row, or a detail row with no owning asset
        body.append(row)
    return headers, body


def _pick(headers: list[str], row: tuple, *names: str) -> str:
    """First non-empty value among the given column names.

    Header names repeat in these sheets (``IP`` and ``Name`` appear under both
    the asset block and the access block), so the FIRST match wins — the asset
    block always comes first.
    """
    for name in names:
        for i, header in enumerate(headers):
            if header.lower() != name.lower():
                continue
            if header.lower() in _FORBIDDEN:
                continue
            if i < len(row):
                value = _clean(row[i])
                if value:
                    return value
    return ""


def _to_int(value: str) -> int | None:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _disk_totals(wb) -> dict[str, float]:
    """Total provisioned disk in GB per asset name, from Disk Information."""
    if "Disk Information" not in wb.sheetnames:
        return {}
    headers, body = _rows(wb["Disk Information"], _DETAIL_HEADER_ROW)
    totals: dict[str, float] = defaultdict(float)
    for row in body:
        name = _pick(headers, row, "*Name", "Name")
        size = _pick(headers, row, "*Capacity(GB)", "Capacity(GB)")
        if not name:
            continue
        try:
            totals[name] += float(size)
        except (TypeError, ValueError):
            continue
    return dict(totals)


def _components(wb) -> dict[str, list[str]]:
    """Application components deployed on each asset."""
    if "Component Information" not in wb.sheetnames:
        return {}
    headers, body = _rows(wb["Component Information"], _DETAIL_HEADER_ROW)
    out: dict[str, list[str]] = defaultdict(list)
    for row in body:
        name = _pick(headers, row, "*Name", "Name")
        component = _pick(headers, row, "Component Type", "*Application unit")
        if name and component and component not in out[name]:
            out[name].append(component)
    return dict(out)


def _extra_ips(wb) -> dict[str, list[str]]:
    """Every IP each asset holds, from VLAN Information."""
    if "VLAN Information" not in wb.sheetnames:
        return {}
    headers, body = _rows(wb["VLAN Information"], _DETAIL_HEADER_ROW)
    out: dict[str, list[str]] = defaultdict(list)
    for row in body:
        name = _pick(headers, row, "*Name", "Name")
        ip = _pick(headers, row, "*IP address", "IP address")
        if name and ip and ip not in out[name]:
            out[name].append(ip)
    return dict(out)


def _app_units(wb) -> dict[str, str]:
    """Application unit per asset, from the IP Information summary sheet."""
    if "IP Information" not in wb.sheetnames:
        return {}
    headers, body = _rows(wb["IP Information"], 1)
    out: dict[str, str] = {}
    for row in body:
        name = _pick(headers, row, "Name")
        unit = _pick(headers, row, "Application unit", "Type")
        if name and unit:
            out.setdefault(name, unit)
    return out


def _host_record(
    headers: list[str],
    row: tuple,
    *,
    kind: str,
    disks: dict[str, float],
    components: dict[str, list[str]],
    extra_ips: dict[str, list[str]],
    app_units: dict[str, str],
    now: datetime,
) -> dict | None:
    """Turn one asset row into a normalized host dict."""
    name = _pick(headers, row, "*Name", "Name", "Host Name")
    if not name:
        return None

    site = _pick(headers, row, "Site name", "Subnet Name")
    ip = _pick(headers, row, "IP", "BMC IP", "MM1IP", "Management IP")
    cores = _to_int(_pick(headers, row, "CPU"))
    mem_mb = _to_int(_pick(headers, row, "Memory(MB)"))

    metrics: dict = {}
    if cores:
        metrics["cores"] = cores
    if mem_mb:
        metrics["mem_total_gb"] = round(mem_mb / 1024, 1)
    if name in disks:
        metrics["disk_total_gb"] = round(disks[name], 1)

    # The application unit (AG, BCS, BTS…) is the grouping an operator actually
    # thinks in; the site is the fallback when the workbook has no unit.
    group = app_units.get(name) or _pick(headers, row, "Application type") or site

    payload = {
        "kind": kind,
        "site": site,
        "os_type": _pick(headers, row, "OS Type"),
        "os_version": _pick(headers, row, "AssetVersion"),
        "vendor": _pick(headers, row, "Vendor"),
        "rack_location": _pick(headers, row, "Location"),
        "model": _pick(headers, row, "Model No.", "*Device type"),
        "serial": _pick(headers, row, "Sequence No."),
        "product": _pick(headers, row, "Product"),
        "deployment_complete": _pick(headers, row, "Deployment complete"),
        "components": components.get(name, []),
        "all_ips": extra_ips.get(name, []),
    }
    return {
        # Site-qualified: the same name can exist at the PR and DR sites.
        "external_id": f"{name}@{site}" if site else name,
        "hostname": name,
        "ip": ip or None,
        # Inventory tells us what exists, never whether it is running.
        "status": HostStatus.unknown,
        "group_name": group or None,
        "last_seen": now,
        "metrics": metrics,
        "raw_payload": {k: v for k, v in payload.items() if v},
    }


def parse_workbook(path: str | Path) -> HuaweiInventory:
    """Read a Digital View asset export into normalized host dicts."""
    from openpyxl import load_workbook

    p = Path(path)
    inventory = HuaweiInventory()
    try:
        stat = p.stat()
        inventory.exported_at = datetime.fromtimestamp(stat.st_mtime, timezone.utc)
    except OSError as exc:
        logger.warning("huawei asset file unreadable (%s): %s", p, exc)
        return inventory

    wb = load_workbook(p, data_only=True, read_only=False)
    try:
        disks = _disk_totals(wb)
        components = _components(wb)
        extra_ips = _extra_ips(wb)
        app_units = _app_units(wb)
        now = datetime.now(timezone.utc)

        seen: set[str] = set()
        for sheets, kind_default in ((_HOST_SHEETS, "host"), (_DEVICE_SHEETS, "device")):
            for sheet, header_row in sheets.items():
                if sheet not in wb.sheetnames:
                    continue
                headers, body = _rows(wb[sheet], header_row)
                added = 0
                for row in body:
                    record = _host_record(
                        headers, row,
                        kind=sheet if kind_default == "device" else sheet,
                        disks=disks, components=components,
                        extra_ips=extra_ips, app_units=app_units, now=now,
                    )
                    if record is None:
                        inventory.skipped += 1
                        continue
                    if record["external_id"] in seen:
                        inventory.skipped += 1
                        continue
                    seen.add(record["external_id"])
                    inventory.hosts.append(record)
                    added += 1
                inventory.sheets_read[sheet] = added
    finally:
        wb.close()

    logger.info(
        "huawei: parsed %d asset(s) from %s (%s)",
        inventory.count, p.name,
        ", ".join(f"{k}={v}" for k, v in inventory.sheets_read.items()),
    )
    return inventory


def load_into_db(db, instance: str, path: str | Path) -> HuaweiInventory:
    """Parse the workbook and upsert its assets as Huawei hosts."""
    from app.normalizer import upsert_hosts

    inventory = parse_workbook(path)
    if inventory.hosts:
        upsert_hosts(db, SourcePlatform.huawei, inventory.hosts, instance)
    return inventory
